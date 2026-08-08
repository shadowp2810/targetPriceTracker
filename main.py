"""
Stock Price Target Tracker — Main Entry Point

Fetches analyst 12-month price targets from yfinance, FMP, and Alpha Vantage,
computes upside / divergence / weekly deltas, and writes a self-contained
HTML dashboard to docs/index.html for GitHub Pages.

Usage:
    python main.py
    python main.py --tickers AAPL MSFT NVDA
    python main.py --excel   # also write reports/targets_YYYY-MM-DD.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, date
from pathlib import Path
from typing import Optional
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from universe import get_universe
from fetcher import fetch_all
from analyzer import analyze_all, build_week_baseline, summary_stats
from exporter_html import write_html

# ---------------------------------------------------------------------------
# openpyxl styles — module level only (never instantiate inside loops)
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
GREEN_FONT = Font(color="375623", bold=True)
RED_FONT = Font(color="9C0006", bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
THIN = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WEEKLY_BASELINE_DAYS = 7
SNAPSHOT_PATH = Path("reports/latest_targets.json")
DOCS_HTML_PATH = Path("docs/index.html")


def load_snapshot(path: Path) -> Optional[dict]:
    if not path.exists():
        return None
    try:
        with open(path, "r") as f:
            return json.load(f)
    except Exception as e:
        print(f"[WARN] Could not load snapshot {path}: {e}")
        return None


def snapshot_age_days(snap: Optional[dict], today: date) -> Optional[int]:
    if not snap or not snap.get("generated"):
        return None
    try:
        return (today - datetime.fromisoformat(snap["generated"]).date()).days
    except ValueError:
        return None


def resolve_week_baseline(snap: Optional[dict], today: date) -> tuple[dict[str, dict], Optional[str], Optional[int]]:
    """
    Return (baseline_by_ticker, baseline_generated, baseline_age_days).

    Prefer an embedded week_baseline that is ~7d old. If the whole snapshot
    itself is >= 7d old and no baseline exists yet, use the snapshot tickers.
    """
    if not snap:
        return {}, None, None

    wb = snap.get("week_baseline") or {}
    wb_generated = wb.get("generated")
    by_ticker = wb.get("by_ticker") or {}

    if by_ticker and wb_generated:
        try:
            age = (today - datetime.fromisoformat(wb_generated).date()).days
            return by_ticker, wb_generated, age
        except ValueError:
            pass

    # Bootstrap: if snapshot is old enough, treat its tickers as the baseline
    age = snapshot_age_days(snap, today)
    if age is not None and age >= WEEKLY_BASELINE_DAYS:
        legacy = {}
        for row in snap.get("tickers") or []:
            if isinstance(row, dict) and row.get("ticker"):
                legacy[row["ticker"]] = {
                    "yf_target_mean": row.get("yf_target_mean"),
                    "fmp_latest_target": row.get("fmp_latest_target"),
                    "av_target": row.get("av_target"),
                    "upside_pct": row.get("upside_pct"),
                    "primary_target": row.get("primary_target"),
                }
        return legacy, snap.get("generated"), age

    return by_ticker, wb_generated, None


def should_refresh_week_baseline(snap: Optional[dict], today: date) -> bool:
    """Refresh embedded week_baseline when missing or >= 7 days old."""
    if not snap:
        return True
    wb = snap.get("week_baseline") or {}
    gen = wb.get("generated")
    if not gen or not wb.get("by_ticker"):
        return True
    try:
        age = (today - datetime.fromisoformat(gen).date()).days
        return age >= WEEKLY_BASELINE_DAYS
    except ValueError:
        return True


def write_excel(analyzed: list[dict], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Targets"

    headers = [
        "Rank", "Ticker", "Name", "Price", "Primary Target", "Source",
        "Upside %", "YF Mean", "YF Median", "FMP Latest", "AV Target",
        "Divergence %", "Above Target", "YF Δ Week", "FMP Δ Week", "AV Δ Week",
        "P/E TTM", "Fwd P/E", "Market Cap", "Avg Volume", "52W High", "52W Low",
        "Beta", "Rec Key", "N Analysts",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for r, row in enumerate(analyzed, 2):
        values = [
            row.get("rank"),
            row.get("ticker"),
            row.get("name"),
            row.get("current_price"),
            row.get("primary_target"),
            row.get("primary_target_source"),
            row.get("upside_pct"),
            row.get("yf_target_mean"),
            row.get("yf_target_median"),
            row.get("fmp_latest_target"),
            row.get("av_target"),
            row.get("divergence_pct"),
            "YES" if row.get("above_target") else "",
            row.get("week_delta_yf"),
            row.get("week_delta_fmp"),
            row.get("week_delta_av"),
            row.get("trailing_pe"),
            row.get("forward_pe"),
            row.get("market_cap"),
            row.get("avg_volume"),
            row.get("fifty_two_week_high"),
            row.get("fifty_two_week_low"),
            row.get("beta"),
            row.get("recommendation_key"),
            row.get("n_analysts"),
        ]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = CENTER if c != 3 else LEFT
            cell.border = THIN_BORDER

            col_name = headers[c - 1]
            if col_name == "Upside %" and val is not None:
                cell.number_format = '0.00"%"'
                if val > 0:
                    cell.fill = GREEN_FILL
                    cell.font = GREEN_FONT
                elif val < 0:
                    cell.fill = RED_FILL
                    cell.font = RED_FONT
            elif col_name == "Above Target" and val == "YES":
                cell.fill = YELLOW_FILL
            elif col_name == "Divergence %" and row.get("divergence_flag"):
                cell.fill = YELLOW_FILL
            elif col_name in ("YF Δ Week", "FMP Δ Week", "AV Δ Week") and val is not None:
                if val > 0:
                    cell.fill = GREEN_FILL
                elif val < 0:
                    cell.fill = RED_FILL

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.column_dimensions["C"].width = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Stock Price Target Tracker")
    parser.add_argument("--tickers", nargs="+", help="Subset of tickers (for local testing)")
    parser.add_argument("--excel", action="store_true", help="Also write an Excel export")
    parser.add_argument("--av-limit", type=int, default=25, help="Alpha Vantage daily refresh limit")
    args = parser.parse_args()

    now = datetime.now(ZoneInfo("America/New_York"))
    iso_timestamp = now.isoformat()
    display_timestamp = now.strftime("%Y-%m-%d %H:%M ET")
    today = now.date()

    tickers = args.tickers if args.tickers else get_universe()
    print(f"Universe: {len(tickers)} tickers")
    print(f"Generated: {display_timestamp}\n")

    reports_dir = Path("reports")
    reports_dir.mkdir(parents=True, exist_ok=True)
    Path("docs").mkdir(parents=True, exist_ok=True)

    prev = load_snapshot(SNAPSHOT_PATH)
    week_baseline, wb_generated, wb_age = resolve_week_baseline(prev, today)
    if wb_generated:
        print(f"Week baseline: {wb_generated[:10]} ({wb_age}d old, {len(week_baseline)} tickers)")
    else:
        print("Week baseline: none (first run or too fresh)")

    print("\n--- Fetching ---")
    fetch_results, av_cache = fetch_all(
        tickers,
        prev_snapshot=prev,
        av_daily_limit=args.av_limit,
    )

    print("\n--- Analyzing ---")
    analyzed = analyze_all(fetch_results, week_baseline_by_ticker=week_baseline)
    stats = summary_stats(analyzed)
    print(
        f"Tickers: {stats['n_tickers']} | "
        f"avg upside: {stats['avg_upside_pct']}% | "
        f"above target: {stats['n_above_target']} | "
        f"divergence flags: {stats['n_divergence_flag']} | "
        f"AV coverage: {stats['n_with_av']}"
    )

    # Persist / refresh week baseline
    if should_refresh_week_baseline(prev, today) and now.weekday() < 5:
        new_week_baseline = {
            "generated": iso_timestamp,
            "by_ticker": build_week_baseline(analyzed),
        }
        print(f"Week baseline refreshed ({WEEKLY_BASELINE_DAYS}d cycle)")
    else:
        new_week_baseline = (prev or {}).get("week_baseline") or {
            "generated": iso_timestamp,
            "by_ticker": build_week_baseline(analyzed),
        }
        print("Week baseline kept")

    snapshot = {
        "generated": iso_timestamp,
        "tickers": analyzed,
        "week_baseline": new_week_baseline,
        "av_cache": av_cache,
        "stats": stats,
    }

    with open(SNAPSHOT_PATH, "w") as f:
        json.dump(snapshot, f, default=str)
    print(f"Snapshot saved: {SNAPSHOT_PATH}")

    snapshot_info = {
        "week_baseline_generated": wb_generated,
        "week_baseline_age_days": wb_age,
        "av_coverage": stats["n_with_av"],
        "av_limit": args.av_limit,
    }
    write_html(
        analyzed,
        DOCS_HTML_PATH,
        display_timestamp,
        iso_timestamp=iso_timestamp,
        stats=stats,
        snapshot_info=snapshot_info,
    )
    print(f"HTML dashboard: {DOCS_HTML_PATH}")

    if args.excel:
        xlsx = reports_dir / f"targets_{today.isoformat()}.xlsx"
        write_excel(analyzed, xlsx)
        print(f"Excel export: {xlsx}")

    print("\nDone.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit(130)
